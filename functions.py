import pandas as pd
import re
from openpyxl import load_workbook


def parse_17():
    xlsx = 'SDVXの難易度表をスプレッドシートで作る.xlsx'
    data = pd.read_excel(xlsx, sheet_name='Lv17', header=1)

    book = load_workbook(xlsx)
    worksheet = book.get_sheet_by_name('Lv17')
    yellow, blue, red = [], [], []
    for i in range(1, worksheet.max_row):
        for j in range(1, worksheet.max_column):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            if fill.start_color.rgb == 'FFFFE599':
                yellow.append(ce.value)
            elif fill.start_color.rgb == 'FFD0E0E3':
                blue.append(ce.value)
            elif fill.start_color.rgb == 'FFF4CCCC':
                red.append(ce.value)

    info_list = []
    for column in data.columns[:-1]:
        titles = data[column].dropna()
        if column in ['A+', 'A', 'B', 'C', 'D']:
            titles = titles[:-1]
        for ori_title in titles:
            if ori_title in yellow:
                class_ = '地力'
            elif ori_title in blue:
                class_ = '鍵盤'
            elif ori_title in red:
                class_ = 'つまみ'
            else:
                class_ = ''
            if ori_title[0] == '【' and ori_title[-1] == '】':
                ori_title = ori_title[1:-1]
                syokenn = 1
            else:
                syokenn = 0
            if ori_title[0] == '※':
                ori_title = ori_title[1:]
                kojinnsa = 1
            else:
                kojinnsa = 0
            if column == '超個人差':
                flag = re.findall(r'\(\d{2}.{0,2000}\)', ori_title)[0][1:-1]
                ori_title = ori_title.split(flag)[0][:-1]
                grade = flag
                kojinnsa = 1
            else:
                grade = column

            title = ori_title
            level = 17

            info_dict = {'title': title, 'level': level, 'grade': grade, 'class': class_, '个人差': kojinnsa,
                         '初見殺し': syokenn}
            info_list.append(info_dict)

    return info_list


def parse_18():
    xlsx = 'SDVXの難易度表をスプレッドシートで作る.xlsx'

    data = pd.read_excel(xlsx, sheet_name='Lv18', header=1)

    book = load_workbook(xlsx)
    worksheet = book.get_sheet_by_name('Lv18')
    yellow, blue, red = [], [], []
    for i in range(1, worksheet.max_row):
        for j in range(1, worksheet.max_column):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            if fill.start_color.rgb == 'FFFFE599':
                yellow.append(ce.value)
            elif fill.start_color.rgb == 'FFD0E0E3':
                blue.append(ce.value)
            elif fill.start_color.rgb == 'FFF4CCCC':
                red.append(ce.value)

    info_list = []
    flag = data[data.SS == 'S'].index[0]
    data['S'] = data['SS'].shift(-flag - 1)
    data['SS'] = data['SS'][:flag]
    data.insert(1, 'S', data.pop('S'))

    for column in data.columns[:-1]:
        titles = data[column].dropna()
        if column in ['超個人差', 'A', 'B', 'C', 'D', 'E']:
            titles = titles[:-1]

        for ori_title in titles:
            if ori_title in yellow:
                class_ = '地力'
            elif ori_title in blue:
                class_ = '鍵盤'
            elif ori_title in red:
                class_ = 'つまみ'
            else:
                class_ = ''
            if ori_title[0] == '【' and ori_title[-1] == '】':
                ori_title = ori_title[1:-1]
                syokenn = 1
            else:
                syokenn = 0
            if ori_title[0] == '※':
                ori_title = ori_title[1:]
                kojinnsa = 1
            else:
                kojinnsa = 0
            if column == '超個人差':
                flag = re.findall(r'\(\d{2}.{0,2000}\)', ori_title)[0][1:-1]
                ori_title = ori_title.split(flag)[0][:-1]
                grade = flag
                kojinnsa = 1
            else:
                grade = column

            title = ori_title
            level = 18

            info_dict = {'title': title, 'level': level, 'grade': grade, 'class': class_, '个人差': kojinnsa,
                         '初見殺し': syokenn}
            info_list.append(info_dict)

    return info_list


def parse_16():
    xlsx = 'SDVXの難易度表をスプレッドシートで作る.xlsx'

    data = pd.read_excel(xlsx, sheet_name='Lv16\xa0', header=1)

    book = load_workbook(xlsx)
    worksheet = book.get_sheet_by_name('Lv16\xa0')
    yellow, blue, red = [], [], []
    for i in range(1, worksheet.max_row):
        for j in range(1, worksheet.max_column):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            if fill.start_color.rgb == 'FFFFE599':
                yellow.append(ce.value)
            elif fill.start_color.rgb == 'FFD0E0E3':
                blue.append(ce.value)
            elif fill.start_color.rgb == 'FFF4CCCC':
                red.append(ce.value)

    info_list = []
    data = data.drop(['Unnamed: 0', '未分類', '変更議論中', 'Unnamed: 12', 'Unnamed: 13', 'Unnamed: 14'], axis=1)
    flag = data[data['16詐称(16S)'] == '初代LV15(現16中～16強)'].index[0]
    data['初代LV15'] = data['16詐称(16S)'].shift(-flag - 1)
    data['16詐称(16S)'] = data['16詐称(16S)'][:flag]
    data.columns = ['16S', '16A～16B_1', '16A～16B_2', '16C～16D_1', '16C～16D_2', '16E～16F_1', '16E～16F_2', '16F', '超個人差',
                    '16A～16B_3']

    for column in data.columns:
        titles = data[column].dropna()
        if column in ['16A～16B_1', '16A～16B_2', '16C～16D_1', '16C～16D_2']:
            titles = titles[:-1]

        for ori_title in titles:
            if ori_title in yellow:
                class_ = '地力'
            elif ori_title in blue:
                class_ = '鍵盤'
            elif ori_title in red:
                class_ = 'つまみ'
            else:
                class_ = ''
            if ori_title[0] == '【' and ori_title[-1] == '】':
                ori_title = ori_title[1:-1]
                syokenn = 1
            else:
                syokenn = 0
            if ori_title[0] == '※':
                ori_title = ori_title[1:]
                kojinnsa = 1
            else:
                kojinnsa = 0
            if column == '超個人差':
                kojinnsa = 1

            grade = column.split('_')[0] if '_' in column else column
            title = ori_title
            if title == 'Juggler`s Maddness[EXH]\n':
                title = 'Juggler`s Maddness[EXH]'
            level = 16

            info_dict = {'title': title, 'level': level, 'grade': grade, 'class': class_, '个人差': kojinnsa,
                         '初見殺し': syokenn}
            info_list.append(info_dict)

    return info_list


def parse_19():
    xlsx = 'C:\\Users\\shzeng\\Downloads\\git\\sdvx_list\\SDVXの難易度表をスプレッドシートで作る.xlsx'

    data = pd.read_excel(xlsx, sheet_name='Lv19〜', header=1)

    book = load_workbook(xlsx)
    worksheet = book.get_sheet_by_name('Lv19〜')
    yellow, blue, red = [], [], []
    for i in range(1, worksheet.max_row):
        for j in range(1, worksheet.max_column):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            if fill.start_color.rgb == 'FFFFE599':
                yellow.append(ce.value)
            elif fill.start_color.rgb == 'FFD0E0E3':
                blue.append(ce.value)
            elif fill.start_color.rgb == 'FFF4CCCC':
                red.append(ce.value)

    info_list = []
    data = data.drop([9, 10], axis=0).reset_index(drop=True)
    flag = data[data['S'] == 'S'].index[0]
    data['19S'] = data['S'].shift(-flag - 1)
    data['20S'] = data['S'][:flag]
    flag = data[data['A+'] == 'A+'].index[0]
    data['19A+'] = data['A+'].shift(-flag - 1)
    data['20A+'] = data['A+'][:flag]
    flag = data[data['A'] == 'A'].index[0]
    data['19A'] = data['A'].shift(-flag - 1)
    data['20A'] = data['A'][:flag]
    flag = data[data['B'] == 'B'].index[0]
    data['19B'] = data['B'].shift(-flag - 1)
    data['20B'] = data['B'][:flag]
    flag = data[data['Unnamed: 5'] == 'C'].index[0]
    data['19C'] = data['Unnamed: 5'].shift(-flag - 1)
    flag = data[data['Unnamed: 6'] == 'D'].index[0]
    data['19D'] = data['Unnamed: 6'].shift(-flag - 1)
    flag = data[data['Unnamed: 7'] == 'E'].index[0]
    data['19E'] = data['Unnamed: 7'].shift(-flag - 1)
    flag = data[data['Unnamed: 8'] == 'F'].index[0]
    data['19F'] = data['Unnamed: 8'].shift(-flag - 1)
    flag = data[data['Unnamed: 9'] == '超個人差'].index[0]
    data['19超個人差'] = data['Unnamed: 9'].shift(-flag - 1)
    data = data.drop(['Lv20', 'S', 'A+', 'A', 'B', 'Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9',
                      'Unnamed: 10'], axis=1)

    for column in data.columns:
        titles = data[column].dropna()

        for ori_title in titles:
            if ori_title in yellow:
                class_ = '地力'
            elif ori_title in blue:
                class_ = '鍵盤'
            elif ori_title in red:
                class_ = 'つまみ'
            else:
                class_ = ''
            if ori_title[0] == '【' and ori_title[-1] == '】':
                ori_title = ori_title[1:-1]
                syokenn = 1
            else:
                syokenn = 0
            if ori_title[0] == '※':
                ori_title = ori_title[1:]
                kojinnsa = 1
            else:
                kojinnsa = 0
            if column == '19超個人差':
                flag = re.findall(r'\(\d{2}.{0,2000}', ori_title)[0][1:]
                ori_title = ori_title.split(flag)[0][:-1]
                grade = flag
                kojinnsa = 1
            else:
                grade = column[2:]

            title = ori_title
            level = column[:2]

            info_dict = {'title': title, 'level': level, 'grade': grade, 'class': class_, '个人差': kojinnsa,
                         '初見殺し': syokenn}
            info_list.append(info_dict)

    return info_list
